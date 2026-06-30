from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, g)
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import datetime, os
import config_loader as cfg
from database import (get_db, init_db, sync_insert, sync_update,
                      sync_delete, pull_from_cloud, push_all_to_cloud)
import supabase_api as supa

app = Flask(__name__)
app.secret_key = cfg.get('SECRET_KEY', 'dgvcl_fallback_key')

# ── Auth helpers ───────────────────────────────────────────────────────────────

PAGES = ['dashboard','divisions','parties','estimates','reports','users','roles']

def get_permissions(role_id):
    db = get_db()
    rows = db.execute("SELECT * FROM dgvcl_permissions WHERE role_id=?", (role_id,)).fetchall()
    db.close()
    return {r['page']: {'view': r['can_view'], 'edit': r['can_edit'], 'delete': r['can_delete']} for r in rows}

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

def perm(page, action='view'):
    def decorator(f):
        @wraps(f)
        def dec(*a, **kw):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            p = g.permissions
            if page not in p or not p[page].get(action, 0):
                flash('You do not have permission to perform this action.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return dec
    return decorator

@app.before_request
def load_user():
    g.user = None
    g.permissions = {}
    if 'user_id' in session:
        db = get_db()
        g.user = db.execute(
            "SELECT u.*, r.name as role_name FROM dgvcl_users u LEFT JOIN dgvcl_roles r ON r.id=u.role_id WHERE u.id=?",
            (session['user_id'],)).fetchone()
        db.close()
        if g.user:
            g.permissions = get_permissions(session.get('role_id', 2))

@app.context_processor
def inject_globals():
    return {
        'app_title': cfg.get('APP_TITLE', 'DGVCL Estimate Portal'),
        'cloud_enabled': supa.is_enabled(),
        'db_mode': 'Supabase ☁' if supa.is_enabled() else 'Local SQLite 💾',
        'now': datetime.datetime.now(),
        'g_permissions': g.permissions if hasattr(g, 'permissions') else {},
    }

# ── Auth ───────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        db   = get_db()
        user = db.execute(
            "SELECT u.*, r.name as role_name FROM dgvcl_users u LEFT JOIN dgvcl_roles r ON r.id=u.role_id WHERE u.username=? AND u.active=1",
            (request.form.get('username', '').strip(),)).fetchone()
        db.close()
        if user and check_password_hash(user['password'], request.form.get('password', '')):
            session.update({'user_id': user['id'], 'username': user['username'],
                            'full_name': user['full_name'], 'role_id': user['role_id'],
                            'role_name': user['role_name']})
            flash(f"Welcome, {user['full_name'] or user['username']}!", 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))

# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@perm('dashboard')
def dashboard():
    db    = get_db()
    today = datetime.date.today()

    # Period filter — supports preset periods OR custom from/to
    period      = request.args.get('period', 'month')
    custom_from = request.args.get('from', '').strip()
    custom_to   = request.args.get('to', '').strip()

    if period == 'custom' and custom_from and custom_to:
        date_from = custom_from
        date_to   = custom_to
    else:
        if period == 'today':
            date_from = today.isoformat()
        elif period == 'week':
            date_from = (today - datetime.timedelta(days=today.weekday())).isoformat()
        elif period == 'month':
            date_from = today.replace(day=1).isoformat()
        elif period == '6month':
            date_from = (today - datetime.timedelta(days=183)).isoformat()
        elif period == 'year':
            date_from = today.replace(month=1, day=1).isoformat()
        else:  # 'all'
            date_from = '2000-01-01'
        date_to = today.isoformat()
        custom_from = date_from
        custom_to   = date_to

    stats = {
        'total_estimates': db.execute("SELECT COUNT(*) FROM dgvcl_estimates WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0],
        'avg_security':    db.execute("SELECT COALESCE(AVG(security_deposit / NULLIF(kw_amount,0)),0) FROM dgvcl_estimates WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0],
        'avg_fixed':       db.execute("SELECT COALESCE(AVG(fixed_charges / NULLIF(kw_amount,0)),0) FROM dgvcl_estimates WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0],
        'avg_getco':       db.execute("SELECT COALESCE(AVG(getco_charge / NULLIF(kw_amount,0)),0) FROM dgvcl_estimates WHERE date>=? AND date<=?", (date_from, date_to)).fetchone()[0],
    }

    recent = db.execute("""
        SELECT e.*, p.name as party_name, d.name as division_name
        FROM dgvcl_estimates e
        JOIN dgvcl_parties p ON p.id = e.party_id
        LEFT JOIN dgvcl_divisions d ON d.id = e.division_id
        WHERE e.date>=? AND e.date<=?
        ORDER BY e.created_at DESC LIMIT 10
    """, (date_from, date_to)).fetchall()

    # Chart data — monthly AVG per-KVA rate (sec/fixed/getco only)
    chart_rows = db.execute("""
        SELECT strftime('%m', date) as mon,
               COALESCE(AVG(security_deposit / NULLIF(kw_amount,0)),0) as sec_dep,
               COALESCE(AVG(fixed_charges    / NULLIF(kw_amount,0)),0) as fixed,
               COALESCE(AVG(getco_charge     / NULLIF(kw_amount,0)),0) as getco
        FROM dgvcl_estimates
        WHERE strftime('%Y', date) = ?
        GROUP BY mon ORDER BY mon
    """, (str(today.year),)).fetchall()

    months_map = {r['mon']: r for r in chart_rows}
    chart = []
    for i, mn in enumerate(['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']):
        key = f"{i+1:02d}"
        r   = months_map.get(key)
        chart.append({'label': mn,
                      'sec_dep': round(r['sec_dep'] if r else 0, 2),
                      'fixed':   round(r['fixed']   if r else 0, 2),
                      'getco':   round(r['getco']   if r else 0, 2)})

    div_stats = [dict(r) for r in db.execute("""
        SELECT d.name, d.location, COUNT(e.id) as est_count,
               COALESCE(SUM(e.security_deposit+e.fixed_charges+e.getco_charge+e.feeder_charge+e.agreement_charges),0) as grand_total
        FROM dgvcl_divisions d
        LEFT JOIN dgvcl_estimates e ON e.division_id = d.id AND e.date>=? AND e.date<=?
        GROUP BY d.id, d.name, d.location
    """, (date_from, date_to)).fetchall()]

    db.close()
    period_labels = {'today':'Today','week':'This Week','month':'This Month',
                     '6month':'Last 6 Months','year':'This Year','all':'All Time','custom':'Custom Range'}
    plabel = period_labels.get(period,'This Month')
    if period == 'custom' and custom_from and custom_to:
        plabel = f"{custom_from} → {custom_to}"
    return render_template('dashboard.html', stats=stats, recent=recent,
                           chart=chart, div_stats=div_stats, period=period,
                           period_label=plabel, custom_from=custom_from, custom_to=custom_to)

# ── Divisions ──────────────────────────────────────────────────────────────────

@app.route('/divisions')
@perm('divisions')
def divisions():
    db   = get_db()
    divs = db.execute("""
        SELECT d.id, d.name, d.location, COUNT(e.id) as est_count,
               COALESCE(SUM(e.security_deposit),0) as total_sec,
               COALESCE(SUM(e.security_deposit+e.fixed_charges+e.getco_charge+e.feeder_charge+e.agreement_charges),0) as grand_total
        FROM dgvcl_divisions d
        LEFT JOIN dgvcl_estimates e ON e.division_id = d.id
        GROUP BY d.id, d.name, d.location ORDER BY d.id
    """).fetchall()
    db.close()
    return render_template('divisions.html', divisions=divs)

@app.route('/divisions/add', methods=['POST'])
@perm('divisions','edit')
def division_add():
    name = request.form.get('name','').strip()
    location = request.form.get('location','').strip()
    if not name:
        flash('Division name is required.', 'danger')
        return redirect(url_for('divisions'))
    db = get_db()
    try:
        cur = db.execute("INSERT INTO dgvcl_divisions (name,location) VALUES (?,?)", (name, location))
        db.commit()
        sync_insert('dgvcl_divisions', {'id': cur.lastrowid, 'name': name, 'location': location})
        flash('Division added.', 'success')
    except Exception:
        flash('Division name already exists.', 'danger')
    db.close()
    return redirect(url_for('divisions'))

@app.route('/divisions/edit/<int:did>', methods=['GET','POST'])
@perm('divisions','edit')
def division_edit(did):
    db = get_db()
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        location = request.form.get('location','').strip()
        if not name:
            flash('Division name is required.', 'danger')
            db.close()
            return redirect(url_for('division_edit', did=did))
        try:
            db.execute("UPDATE dgvcl_divisions SET name=?, location=? WHERE id=?", (name, location, did))
            db.commit()
            sync_update('dgvcl_divisions', did, {'name': name, 'location': location})
            flash('Division updated.', 'success')
            db.close()
            return redirect(url_for('divisions'))
        except Exception:
            flash('Division name already exists.', 'danger')
            db.close()
            return redirect(url_for('division_edit', did=did))
    division = db.execute("SELECT * FROM dgvcl_divisions WHERE id=?", (did,)).fetchone()
    db.close()
    if not division:
        flash('Division not found.', 'danger')
        return redirect(url_for('divisions'))
    return render_template('division_form.html', division=division)

@app.route('/divisions/delete/<int:did>', methods=['POST'])
@perm('divisions','delete')
def division_delete(did):
    db = get_db()
    party_count = db.execute("SELECT COUNT(*) FROM dgvcl_parties WHERE division_id=?", (did,)).fetchone()[0]
    est_count   = db.execute("SELECT COUNT(*) FROM dgvcl_estimates WHERE division_id=?", (did,)).fetchone()[0]
    if party_count or est_count:
        flash(f'Cannot delete — division has {party_count} part(y/ies) and {est_count} estimate(s) linked to it.', 'danger')
    else:
        db.execute("DELETE FROM dgvcl_divisions WHERE id=?", (did,))
        db.commit()
        sync_delete('dgvcl_divisions', did)
        flash('Division deleted.', 'success')
    db.close()
    return redirect(url_for('divisions'))

# ── Parties ────────────────────────────────────────────────────────────────────

@app.route('/parties')
@perm('parties')
def parties():
    db   = get_db()
    rows = db.execute("""
        SELECT p.*, d.name as division_name,
               COUNT(e.id) as est_count,
               COALESCE(SUM(e.security_deposit+e.fixed_charges+e.getco_charge+e.feeder_charge+e.agreement_charges),0) as total_value
        FROM dgvcl_parties p
        LEFT JOIN dgvcl_divisions d ON d.id = p.division_id
        LEFT JOIN dgvcl_estimates e ON e.party_id = p.id
        WHERE p.active=1
        GROUP BY p.id ORDER BY p.name
    """).fetchall()
    divs = db.execute("SELECT * FROM dgvcl_divisions ORDER BY name").fetchall()
    db.close()
    return render_template('parties.html', parties=rows, divisions=divs)

@app.route('/parties/add', methods=['POST'])
@perm('parties','edit')
def party_add():
    f = request.form
    name = f.get('name','').strip()
    if not name:
        flash('Party name is required.', 'danger')
        return redirect(url_for('parties'))
    db = get_db()
    cur = db.execute(
        "INSERT INTO dgvcl_parties (name,contact,email,address,gst,division_id) VALUES (?,?,?,?,?,?)",
        (name, f.get('contact'), f.get('email'), f.get('address'),
         f.get('gst'), f.get('division_id') or None))
    db.commit()
    sync_insert('dgvcl_parties', {'id': cur.lastrowid, 'name': name,
        'contact': f.get('contact'), 'email': f.get('email'),
        'address': f.get('address'), 'gst': f.get('gst'),
        'division_id': f.get('division_id') or None, 'active': 1})
    db.close()
    flash('Party added successfully.', 'success')
    return redirect(url_for('parties'))

@app.route('/parties/edit/<int:pid>', methods=['GET','POST'])
@perm('parties','edit')
def party_edit(pid):
    db = get_db()
    if request.method == 'POST':
        f = request.form
        db.execute("""UPDATE dgvcl_parties SET name=?,contact=?,email=?,address=?,gst=?,division_id=?
                      WHERE id=?""",
                   (f.get('name'), f.get('contact'), f.get('email'),
                    f.get('address'), f.get('gst'), f.get('division_id') or None, pid))
        db.commit()
        sync_update('dgvcl_parties', pid, dict(f))
        db.close()
        flash('Party updated.', 'success')
        return redirect(url_for('parties'))
    party = db.execute("SELECT * FROM dgvcl_parties WHERE id=?", (pid,)).fetchone()
    divs  = db.execute("SELECT * FROM dgvcl_divisions ORDER BY name").fetchall()
    db.close()
    return render_template('party_form.html', party=party, divisions=divs)

@app.route('/parties/delete/<int:pid>', methods=['POST'])
@perm('parties','delete')
def party_delete(pid):
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM dgvcl_estimates WHERE party_id=?", (pid,)).fetchone()[0]
    if count:
        flash(f'Cannot delete — party has {count} estimate(s).', 'danger')
    else:
        db.execute("DELETE FROM dgvcl_parties WHERE id=?", (pid,))
        db.commit()
        sync_delete('dgvcl_parties', pid)
        flash('Party deleted.', 'success')
    db.close()
    return redirect(url_for('parties'))

# ── Estimates ──────────────────────────────────────────────────────────────────

@app.route('/estimates')
@perm('estimates')
def estimates():
    db      = get_db()
    parties = db.execute("SELECT * FROM dgvcl_parties WHERE active=1 ORDER BY name").fetchall()
    divs    = db.execute("SELECT * FROM dgvcl_divisions ORDER BY name").fetchall()
    db.close()
    return render_template('estimates.html', parties=parties, divisions=divs)

@app.route('/estimates/add', methods=['POST'])
@perm('estimates','edit')
def estimate_add():
    f = request.form
    errors = []
    if not f.get('date'):            errors.append('Date is required.')
    if not f.get('party_id'):        errors.append('Party is required.')
    if not f.get('scope_of_work'):   errors.append('Scope of work is required.')
    if not f.get('security_deposit') or float(f.get('security_deposit',0)) <= 0:
        errors.append('Security Deposit is required and must be > 0.')
    if not f.get('fixed_charges') or float(f.get('fixed_charges',0)) <= 0:
        errors.append('Fixed Charges is required and must be > 0.')
    if not f.get('getco_charge') or float(f.get('getco_charge',0)) <= 0:
        errors.append('GETCO Charge is required and must be > 0.')
    if errors:
        for e in errors: flash(e, 'danger')
        return redirect(url_for('estimates'))

    kw   = float(f.get('kw_amount') or 1)
    sec  = round(float(f['security_deposit']) * kw, 2)
    fix  = round(float(f['fixed_charges'])    * kw, 2)
    getco= round(float(f['getco_charge'])     * kw, 2)
    feed = round(float(f.get('feeder_charge') or 0), 2)      # NOT multiplied by KVA
    agr  = round(float(f.get('agreement_charges') or 0), 2)  # NOT multiplied by KVA

    db  = get_db()
    cur = db.execute("""
        INSERT INTO dgvcl_estimates
            (date,party_id,division_id,scope_of_work,kw_amount,security_deposit,
             fixed_charges,getco_charge,feeder_charge,agreement_charges,created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (f['date'], f['party_id'], f.get('division_id') or None, f['scope_of_work'],
          kw, sec, fix, getco, feed, agr, session['user_id']))
    db.commit()
    sync_insert('dgvcl_estimates', {'id': cur.lastrowid, 'date': f['date'],
        'party_id': f['party_id'], 'division_id': f.get('division_id') or None,
        'scope_of_work': f['scope_of_work'], 'kw_amount': kw,
        'security_deposit': sec, 'fixed_charges': fix, 'getco_charge': getco,
        'feeder_charge': feed, 'agreement_charges': agr, 'created_by': session['user_id']})
    db.close()
    flash('Estimate saved successfully.', 'success')
    return redirect(url_for('estimates'))

@app.route('/estimates/edit/<int:eid>', methods=['GET','POST'])
@perm('estimates','edit')
def estimate_edit(eid):
    db = get_db()
    if request.method == 'POST':
        f = request.form
        kw   = float(f.get('kw_amount') or 1)
        sec  = round(float(f['security_deposit']) * kw, 2)
        fix  = round(float(f['fixed_charges'])    * kw, 2)
        getco= round(float(f['getco_charge'])     * kw, 2)
        feed = round(float(f.get('feeder_charge') or 0), 2)      # NOT multiplied by KVA
        agr  = round(float(f.get('agreement_charges') or 0), 2)  # NOT multiplied by KVA
        db.execute("""UPDATE dgvcl_estimates SET date=?,party_id=?,division_id=?,scope_of_work=?,
                      kw_amount=?,security_deposit=?,fixed_charges=?,getco_charge=?,
                      feeder_charge=?,agreement_charges=? WHERE id=?""",
                   (f['date'], f['party_id'], f.get('division_id') or None, f['scope_of_work'],
                    kw, sec, fix, getco, feed, agr, eid))
        db.commit()
        sync_update('dgvcl_estimates', eid, dict(f))
        db.close()
        flash('Estimate updated.', 'success')
        return redirect(url_for('estimates'))
    est     = db.execute("SELECT * FROM dgvcl_estimates WHERE id=?", (eid,)).fetchone()
    parties = db.execute("SELECT * FROM dgvcl_parties WHERE active=1 ORDER BY name").fetchall()
    divs    = db.execute("SELECT * FROM dgvcl_divisions ORDER BY name").fetchall()
    db.close()
    return render_template('estimate_form.html', est=est, parties=parties, divisions=divs)

@app.route('/estimates/delete/<int:eid>', methods=['POST'])
@perm('estimates','delete')
def estimate_delete(eid):
    db = get_db()
    db.execute("DELETE FROM dgvcl_estimates WHERE id=?", (eid,))
    db.commit()
    sync_delete('dgvcl_estimates', eid)
    db.close()
    flash('Estimate deleted.', 'success')
    return redirect(url_for('estimates'))



# ── API — list estimates (for AJAX datatable) ──────────────────────────────────

@app.route('/api/estimates')
@perm('estimates')
def api_estimates():
    db          = get_db()
    division_id = request.args.get('division_id')
    party_id    = request.args.get('party_id')
    from_date   = request.args.get('from')
    to_date     = request.args.get('to')

    q      = """SELECT e.*, p.name as party_name, d.name as division_name,
                       (e.security_deposit+e.fixed_charges+e.getco_charge+
                        e.feeder_charge+e.agreement_charges) as total
                FROM dgvcl_estimates e
                JOIN dgvcl_parties p ON p.id=e.party_id
                LEFT JOIN dgvcl_divisions d ON d.id=e.division_id
                WHERE 1=1"""
    params = []
    if division_id: q += " AND e.division_id=?"; params.append(division_id)
    if party_id:    q += " AND e.party_id=?";    params.append(party_id)
    if from_date:   q += " AND e.date>=?";        params.append(from_date)
    if to_date:     q += " AND e.date<=?";        params.append(to_date)
    q += " ORDER BY e.date DESC"

    rows = db.execute(q, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

# ── Reports ────────────────────────────────────────────────────────────────────

@app.route('/reports')
@perm('reports')
def reports():
    db      = get_db()
    divs    = db.execute("SELECT * FROM dgvcl_divisions ORDER BY name").fetchall()
    parties = db.execute("SELECT * FROM dgvcl_parties WHERE active=1 ORDER BY name").fetchall()
    db.close()
    return render_template('reports.html', divisions=divs, parties=parties)


# ── Export routes ──────────────────────────────────────────────────────────────

@app.route('/reports/export/excel')
@perm('reports')
def export_excel():
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db          = get_db()
    division_id = request.args.get('division_id')
    party_id    = request.args.get('party_id')
    from_date   = request.args.get('from')
    to_date     = request.args.get('to')

    q      = """SELECT e.date, p.name as party_name, d.name as division_name,
                       e.scope_of_work, e.kw_amount,
                       e.security_deposit, e.fixed_charges, e.getco_charge,
                       e.feeder_charge, e.agreement_charges,
                       (e.security_deposit+e.fixed_charges+e.getco_charge+
                        e.feeder_charge+e.agreement_charges) as total
                FROM dgvcl_estimates e
                JOIN dgvcl_parties p ON p.id=e.party_id
                LEFT JOIN dgvcl_divisions d ON d.id=e.division_id
                WHERE 1=1"""
    params = []
    if division_id: q += " AND e.division_id=?"; params.append(division_id)
    if party_id:    q += " AND e.party_id=?";    params.append(party_id)
    if from_date:   q += " AND e.date>=?";        params.append(from_date)
    if to_date:     q += " AND e.date<=?";        params.append(to_date)
    q += " ORDER BY e.date DESC"
    rows = db.execute(q, params).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DGVCL Estimates"

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "DGVCL Estimate Report"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="0D1526")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    ws['A2'] = f"Generated: {datetime.date.today().strftime('%d %B %Y')}"
    ws['A2'].font = Font(size=10, color="7A8DB0")
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Date','Party','Division','Scope of Work','KVA',
               'Security Deposit','Fixed Charges','GETCO Charge',
               'Feeder Charge','Agreement Charges','Total']
    col_widths = [14,25,22,35,8,18,16,16,16,20,18]

    hdr_fill = PatternFill("solid", fgColor="1E6FFF")
    thin = Border(bottom=Side(style='thin', color='1E6FFF'))

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    # Helper — format as Rs. string
    def rs(v): return f"Rs. {int(round(float(v or 0))):,}"

    alt_fill = PatternFill("solid", fgColor="F0F4FF")
    for ri, row in enumerate(rows, 4):
        alt = PatternFill("solid", fgColor="F8FAFF") if ri % 2 == 0 else None
        vals = [row['date'], row['party_name'], row['division_name'] or '—',
                row['scope_of_work'], f"{row['kw_amount']} KVA",
                rs(row['security_deposit']), rs(row['fixed_charges']), rs(row['getco_charge']),
                rs(row['feeder_charge']), rs(row['agreement_charges']), rs(row['total'])]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if alt: cell.fill = alt
            if ci >= 6:
                cell.alignment = Alignment(horizontal='right')
            cell.font = Font(size=10)

    # Totals row
    tr = len(rows) + 4
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, size=10)
    ws.merge_cells(f'A{tr}:E{tr}')
    keys = ['security_deposit','fixed_charges','getco_charge','feeder_charge','agreement_charges','total']
    for ci, key in enumerate(keys, 6):
        cell = ws.cell(row=tr, column=ci,
                       value=rs(sum(float(r[key] or 0) for r in rows)))
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E6F1FB")
        cell.alignment = Alignment(horizontal='right')

    ws.freeze_panes = 'A4'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    fname = f"DGVCL_Estimates_{datetime.date.today()}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/reports/export/pdf')
@perm('reports')
def export_pdf():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    db          = get_db()
    division_id = request.args.get('division_id')
    party_id    = request.args.get('party_id')
    from_date   = request.args.get('from')
    to_date     = request.args.get('to')

    q      = """SELECT e.date, p.name as party_name, d.name as division_name,
                       e.scope_of_work, e.kw_amount,
                       e.security_deposit, e.fixed_charges, e.getco_charge,
                       e.feeder_charge, e.agreement_charges,
                       (e.security_deposit+e.fixed_charges+e.getco_charge+
                        e.feeder_charge+e.agreement_charges) as total
                FROM dgvcl_estimates e
                JOIN dgvcl_parties p ON p.id=e.party_id
                LEFT JOIN dgvcl_divisions d ON d.id=e.division_id
                WHERE 1=1"""
    params = []
    if division_id: q += " AND e.division_id=?"; params.append(division_id)
    if party_id:    q += " AND e.party_id=?";    params.append(party_id)
    if from_date:   q += " AND e.date>=?";        params.append(from_date)
    if to_date:     q += " AND e.date<=?";        params.append(to_date)
    q += " ORDER BY e.date DESC"
    rows = db.execute(q, params).fetchall()
    db.close()

    # Use A4 landscape — compute usable width
    PAGE = landscape(A4)
    usable_w = PAGE[0] - 20*mm  # 10mm margin each side

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)

    # Register Unicode font for rupee symbol support
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _font_registered = False
    for font_path in ['/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                      '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
                      '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                      '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf']:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('UniSans', font_path))
                bold_path = font_path.replace('.ttf','-Bold.ttf')
                if not os.path.exists(bold_path): bold_path = font_path
                pdfmetrics.registerFont(TTFont('UniSans-Bold', bold_path))
                _font_registered = True
                break
            except Exception:
                pass
    FONT_NORMAL = 'UniSans' if _font_registered else 'Helvetica'
    FONT_BOLD   = 'UniSans-Bold' if _font_registered else 'Helvetica-Bold'

    navy   = colors.HexColor('#0D1526')
    blue   = colors.HexColor('#1E6FFF')
    ltblue = colors.HexColor('#EEF4FF')
    altrow = colors.HexColor('#F5F8FF')
    white  = colors.white
    muted  = colors.HexColor('#5A6A8A')
    dark   = colors.HexColor('#1A2035')

    title_style = ParagraphStyle('t', fontSize=15, textColor=white,
                                 fontName=FONT_BOLD,
                                 backColor=navy, alignment=TA_CENTER,
                                 spaceAfter=0, spaceBefore=0, leading=26,
                                 leftPadding=8, rightPadding=8)
    sub_style   = ParagraphStyle('s', fontSize=9, textColor=muted,
                                 alignment=TA_CENTER, spaceAfter=6, leading=14)

    # Format date properly dd-Mon-YYYY and amount with rupee symbol
    def fmt_date(d):
        try:
            import datetime as dt
            return dt.datetime.strptime(str(d), '%Y-%m-%d').strftime('%d-%b-%Y')
        except Exception:
            return str(d)

    def inr(v):
        amt = float(v or 0)
        s = f"{int(round(amt)):,}"
        return f"Rs.{s}"

    # Column proportions out of usable_w (must sum to 1.0)
    # Date  Party  Division  Scope   KVA   Sec    Fixed  GETCO  Feeder  Agr   Total
    props = [0.08, 0.13,     0.10,   0.18, 0.06,  0.09,  0.075, 0.075, 0.07, 0.065, 0.08]
    col_widths = [p * usable_w for p in props]

    hdrs = ['Date','Party','Division','Scope of Work','KVA',
            'Sec. Dep.','Fixed Chg.','GETCO Chg.','Feeder Chg.','Agr. Chg.','Total']
    data = [hdrs]
    for r in rows:
        scope = r['scope_of_work']
        if len(scope) > 35: scope = scope[:34] + '…'
        data.append([
            fmt_date(r['date']),
            r['party_name'],
            (r['division_name'] or '—').split('–')[0].strip(),
            scope,
            f"{r['kw_amount']} KVA",
            inr(r['security_deposit']),
            inr(r['fixed_charges']),
            inr(r['getco_charge']),
            inr(r['feeder_charge']),
            inr(r['agreement_charges']),
            inr(r['total']),
        ])

    # Totals row
    data.append([
        'TOTAL', '', '', '', '',
        inr(sum(float(r['security_deposit'] or 0) for r in rows)),
        inr(sum(float(r['fixed_charges'] or 0) for r in rows)),
        inr(sum(float(r['getco_charge'] or 0) for r in rows)),
        inr(sum(float(r['feeder_charge'] or 0) for r in rows)),
        inr(sum(float(r['agreement_charges'] or 0) for r in rows)),
        inr(sum(float(r['total'] or 0) for r in rows)),
    ])

    last = len(data) - 1

    style_cmds = [
        # Header row
        ('BACKGROUND',   (0,0),  (-1,0),     blue),
        ('TEXTCOLOR',    (0,0),  (-1,0),     white),
        ('FONTNAME',     (0,0),  (-1,0),     FONT_BOLD),
        ('FONTSIZE',     (0,0),  (-1,0),     7.5),
        ('ALIGN',        (0,0),  (-1,0),     'CENTER'),
        ('VALIGN',       (0,0),  (-1,0),     'MIDDLE'),
        ('ROWHEIGHT',    (0,0),  (0,0),      18),
        # Data rows
        ('FONTNAME',     (0,1),  (-1,last-1), FONT_NORMAL),
        ('FONTSIZE',     (0,1),  (-1,last-1), 7.5),
        ('TEXTCOLOR',    (0,1),  (-1,last-1), dark),
        ('VALIGN',       (0,1),  (-1,-1),    'MIDDLE'),
        ('ROWHEIGHT',    (0,1),  (-1,last-1), 16),
        # Align numbers right
        ('ALIGN',        (5,1),  (-1,-1),    'RIGHT'),
        ('ALIGN',        (0,1),  (4,-1),     'LEFT'),
        # Totals row
        ('BACKGROUND',   (0,last),(-1,last), ltblue),
        ('FONTNAME',     (0,last),(-1,last), FONT_BOLD),
        ('FONTSIZE',     (0,last),(-1,last), 7.5),
        ('TEXTCOLOR',    (0,last),(-1,last), dark),
        ('ROWHEIGHT',    (0,last),(0,last),  18),
        # Alternating rows
        *[('BACKGROUND', (0,i),(-1,i), altrow) for i in range(2, last, 2)],
        # Grid
        ('GRID',         (0,0),  (-1,-1),    0.25, colors.HexColor('#C8D4EE')),
        ('LINEBELOW',    (0,0),  (-1,0),     1.0,  blue),
        ('LINEABOVE',    (0,last),(-1,last), 0.8,  blue),
    ]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))

    filter_info = []
    if from_date: filter_info.append(f"From: {fmt_date(from_date)}")
    if to_date:   filter_info.append(f"To: {fmt_date(to_date)}")
    filter_str = "  |  ".join(filter_info) if filter_info else "All records"

    elements = [
        Paragraph("DGVCL Estimate Report", title_style),
        Spacer(1, 3*mm),
        Paragraph(
            f"Generated: {datetime.date.today().strftime('%d %B %Y')}  "
            f"|  {len(rows)} records  |  {filter_str}",
            sub_style),
        tbl,
    ]
    doc.build(elements)
    buf.seek(0)
    from flask import send_file
    fname = f"DGVCL_Estimates_{datetime.date.today()}.pdf"
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True, download_name=fname)

# ── Cloud Settings ─────────────────────────────────────────────────────────────

@app.route('/cloud', methods=['GET','POST'])
@perm('users')
def cloud_settings():
    from config_loader import save_to_xlsx, get_masked_config, config_file_exists
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save_keys':
            save_to_xlsx({
                'SUPABASE_URL':      request.form.get('url','').strip(),
                'SUPABASE_ANON_KEY': request.form.get('key','').strip(),
            })
            supa.reload_keys()
            flash('Keys saved.', 'success')
        elif action == 'test':
            ok, msg = supa.test_connection()
            flash(msg, 'success' if ok else 'danger')
        elif action == 'push':
            results = push_all_to_cloud()
            ok = sum(1 for v in results.values() if v)
            flash(f'Pushed {ok}/{len(results)} tables to Supabase.', 'success' if ok else 'danger')
        elif action == 'pull':
            if pull_from_cloud():
                flash('Pulled latest data from Supabase.', 'success')
            else:
                flash('Pull failed or Supabase not connected.', 'danger')
        elif action == 'get_sql':
            return render_template('cloud_settings.html',
                                   masked=get_masked_config(),
                                   sql=supa.create_tables_sql(),
                                   file_exists=config_file_exists())
        return redirect(url_for('cloud_settings'))

    return render_template('cloud_settings.html',
                           masked=get_masked_config(),
                           sql=None,
                           file_exists=config_file_exists())


# ── Roles ──────────────────────────────────────────────────────────────────────

@app.route('/roles')
@perm('roles')
def roles_list():
    db    = get_db()
    roles = db.execute("""
        SELECT r.*, COUNT(u.id) as user_count
        FROM dgvcl_roles r
        LEFT JOIN dgvcl_users u ON u.role_id = r.id
        GROUP BY r.id ORDER BY r.name
    """).fetchall()
    # Build permissions dict per role for display
    all_perms = {}
    for role in roles:
        rows = db.execute("SELECT * FROM dgvcl_permissions WHERE role_id=?", (role['id'],)).fetchall()
        all_perms[role['id']] = {r['page']: {'view':r['can_view'],'edit':r['can_edit'],'delete':r['can_delete']} for r in rows}
    db.close()
    return render_template('roles.html', roles=roles, all_perms=all_perms, pages=PAGES)

@app.route('/roles/add', methods=['GET','POST'])
@perm('roles','edit')
def role_add():
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        if not name:
            flash('Role name is required.', 'danger')
            return redirect(url_for('roles_list'))
        db = get_db()
        cur = db.execute("INSERT INTO dgvcl_roles (name,description) VALUES (?,?)",
                         (name, request.form.get('description','').strip()))
        role_id = cur.lastrowid
        # Save permissions
        for page in PAGES:
            can_view   = 1 if request.form.get(f'{page}_view')   else 0
            can_edit   = 1 if request.form.get(f'{page}_edit')   else 0
            can_delete = 1 if request.form.get(f'{page}_delete') else 0
            db.execute("INSERT INTO dgvcl_permissions (role_id,page,can_view,can_edit,can_delete) VALUES (?,?,?,?,?)",
                       (role_id, page, can_view, can_edit, can_delete))
        db.commit(); db.close()
        flash('Role created.', 'success')
        return redirect(url_for('roles_list'))
    return render_template('role_form.html', action='Add', role=None, pages=PAGES, perms={})

@app.route('/roles/edit/<int:rid>', methods=['GET','POST'])
@perm('roles','edit')
def role_edit(rid):
    db = get_db()
    if request.method == 'POST':
        db.execute("UPDATE dgvcl_roles SET name=?,description=? WHERE id=?",
                   (request.form.get('name','').strip(),
                    request.form.get('description','').strip(), rid))
        for page in PAGES:
            can_view   = 1 if request.form.get(f'{page}_view')   else 0
            can_edit   = 1 if request.form.get(f'{page}_edit')   else 0
            can_delete = 1 if request.form.get(f'{page}_delete') else 0
            existing = db.execute("SELECT id FROM dgvcl_permissions WHERE role_id=? AND page=?",
                                  (rid, page)).fetchone()
            if existing:
                db.execute("UPDATE dgvcl_permissions SET can_view=?,can_edit=?,can_delete=? WHERE role_id=? AND page=?",
                           (can_view, can_edit, can_delete, rid, page))
            else:
                db.execute("INSERT INTO dgvcl_permissions (role_id,page,can_view,can_edit,can_delete) VALUES (?,?,?,?,?)",
                           (rid, page, can_view, can_edit, can_delete))
        db.commit(); db.close()
        flash('Role updated.', 'success')
        return redirect(url_for('roles_list'))
    role  = db.execute("SELECT * FROM dgvcl_roles WHERE id=?", (rid,)).fetchone()
    perms = {r['page']: {'view': r['can_view'], 'edit': r['can_edit'], 'delete': r['can_delete']}
             for r in db.execute("SELECT * FROM dgvcl_permissions WHERE role_id=?", (rid,)).fetchall()}
    db.close()
    return render_template('role_form.html', action='Edit', role=role, pages=PAGES, perms=perms)

@app.route('/roles/delete/<int:rid>', methods=['POST'])
@perm('roles','delete')
def role_delete(rid):
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM dgvcl_users WHERE role_id=?", (rid,)).fetchone()[0]
    if count:
        flash(f'Cannot delete — {count} user(s) use this role.', 'danger')
    else:
        db.execute("DELETE FROM dgvcl_roles WHERE id=?", (rid,))
        db.commit()
        flash('Role deleted.', 'success')
    db.close()
    return redirect(url_for('roles_list'))

# ── Users ──────────────────────────────────────────────────────────────────────

@app.route('/users')
@perm('users')
def users():
    db    = get_db()
    rows  = db.execute("""
        SELECT u.*, r.name as role_name FROM dgvcl_users u
        LEFT JOIN dgvcl_roles r ON r.id=u.role_id
        ORDER BY u.username
    """).fetchall()
    roles = db.execute("SELECT * FROM dgvcl_roles ORDER BY name").fetchall()
    db.close()
    return render_template('users.html', users=rows, roles=roles)

@app.route('/users/add', methods=['POST'])
@perm('users','edit')
def user_add():
    f    = request.form
    name = f.get('username','').strip()
    pwd  = f.get('password','').strip()
    if not name or not pwd:
        flash('Username and password are required.', 'danger')
        return redirect(url_for('users'))
    db = get_db()
    try:
        db.execute("INSERT INTO dgvcl_users (username,password,full_name,role_id) VALUES (?,?,?,?)",
                   (name, generate_password_hash(pwd), f.get('full_name'), int(f.get('role_id', 2))))
        db.commit()
        flash('User added.', 'success')
    except Exception:
        flash('Username already exists.', 'danger')
    db.close()
    return redirect(url_for('users'))

@app.route('/users/delete/<int:uid>', methods=['POST'])
@perm('users','delete')
def user_delete(uid):
    if uid == session.get('user_id'):
        flash('Cannot delete yourself.', 'danger')
        return redirect(url_for('users'))
    db = get_db()
    db.execute("DELETE FROM dgvcl_users WHERE id=?", (uid,))
    db.commit()
    db.close()
    flash('User deleted.', 'success')
    return redirect(url_for('users'))

# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    # DEBUG=True locally, False on PythonAnywhere (set env var DEBUG=0 there)
    debug = os.environ.get('DEBUG', '1') != '0'
    app.run(debug=debug, host='0.0.0.0', port=5000)
