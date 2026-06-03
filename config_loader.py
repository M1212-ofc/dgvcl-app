"""
config_loader.py  —  reads DGVCL settings from config.xlsx
Edit config.xlsx (Key / Value columns) to set your Supabase credentials.
"""
import os
try:
    import openpyxl
    _HAS_XLSX = True
except ImportError:
    _HAS_XLSX = False

CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config.xlsx')
_cache = {}

DEFAULTS = {
    'SUPABASE_URL':      '',
    'SUPABASE_ANON_KEY': '',
    'SECRET_KEY':        'dgvcl_xK9_change_in_production',
    'APP_TITLE':         'DGVCL Estimate Portal',
}

def _load():
    global _cache
    _cache = dict(DEFAULTS)
    if not _HAS_XLSX or not os.path.exists(CONFIG_FILE):
        return
    try:
        wb = openpyxl.load_workbook(CONFIG_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                _cache[str(row[0]).strip()] = str(row[1]).strip()
    except Exception as e:
        print(f"[Config] failed to load config.xlsx: {e}")

def reload():
    _load()

def get(key, default=''):
    if not _cache:
        _load()
    return _cache.get(key, default)

def config_file_exists():
    return os.path.exists(CONFIG_FILE)

def get_masked_config():
    if not _cache:
        _load()
    out = {}
    for k, v in _cache.items():
        if 'KEY' in k and len(v) > 10:
            out[k] = v[:6] + '…' + v[-4:]
        else:
            out[k] = v
    return out

def save_to_xlsx(data: dict):
    if not _HAS_XLSX:
        return False
    try:
        if os.path.exists(CONFIG_FILE):
            wb = openpyxl.load_workbook(CONFIG_FILE)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Key', 'Value'])
        existing = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                existing[row[0].value] = row[0].row
        for k, v in data.items():
            if k in existing:
                ws.cell(row=existing[k], column=2, value=v)
            else:
                ws.append([k, v])
        wb.save(CONFIG_FILE)
        _load()
        return True
    except Exception as e:
        print(f"[Config] save failed: {e}")
        return False
